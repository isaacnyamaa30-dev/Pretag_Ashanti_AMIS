"""Write a two-period comparison to an .xlsx workbook (Regional / Zones / Districts / Movers)."""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# PRETAG brand: red header, logo-derived status washes
_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="C4161C")
_STATUS_FILL = {
    "GROWING": PatternFill("solid", fgColor="E9EFD7"),
    "DECLINING": PatternFill("solid", fgColor="F3DED9"),
    "STABLE": PatternFill("solid", fgColor="F5E8C6"),
    "NEW": PatternFill("solid", fgColor="ECE6D6"),
}
_COLS = ["name", "previous", "current", "added", "missing",
         "transfers_in", "transfers_out", "net", "growth_pct", "retention_pct", "status"]
_LABELS = ["Name", "Previous", "Current", "Added", "Missing", "Transfers in",
           "Transfers out", "Net change", "Growth %", "Retention %", "Status"]


def _sheet(wb, title, lines, extra_first=None):
    ws = wb.create_sheet(title)
    labels = (["Zone"] if extra_first else []) + _LABELS
    ws.append(labels)
    for c in ws[1]:
        c.font = _HEAD
        c.fill = _HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    for ln in lines:
        prefix = [ln["zone"]] if extra_first else []
        ws.append(prefix + [ln.get(k) for k in _COLS])
        if ln.get("status") in _STATUS_FILL:
            ws.cell(row=ws.max_row, column=ws.max_column).fill = _STATUS_FILL[ln["status"]]
    widths = ([26] if extra_first else []) + [26, 10, 10, 8, 9, 12, 13, 11, 10, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def write(result: dict, path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ov = wb.create_sheet("Overview")
    r = result["regional"]
    ov.append(["PRETAG Ashanti - Membership Comparison"])
    ov["A1"].font = Font(bold=True, size=14)
    ov.append([])
    ov.append(["Previous file", result["previous_file"]])
    ov.append(["Current file", result["current_file"]])
    ov.append([])
    for k, label in [("previous", "Opening membership"), ("current", "Closing membership"),
                     ("added", "Added to current R20"), ("missing", "Missing from current R20"),
                     ("net", "Net change"), ("growth_pct", "Growth %"),
                     ("retention_pct", "Retention %"), ("status", "Regional status")]:
        ov.append([label, r[k]])
    ov.append([])
    for s, n in result["zone_status_counts"].items():
        ov.append([f"{s.title()} zones", n])
    ov.column_dimensions["A"].width = 28
    ov.column_dimensions["B"].width = 24

    _sheet(wb, "Zones", result["zones"])
    _sheet(wb, "Districts", result["districts"], extra_first=True)

    mv = wb.create_sheet("Internal movers")
    mv.append(["Employee no", "Name", "From zone", "To zone", "From district", "To district"])
    for c in mv[1]:
        c.font = _HEAD
        c.fill = _HEAD_FILL
    for m in result["internal_movers"]:
        mv.append([m["employee_no"], m["name"], m["from_zone"], m["to_zone"],
                   m["from_district"], m["to_district"]])
    for i, w in enumerate([14, 28, 20, 20, 24, 24], start=1):
        mv.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    mv.freeze_panes = "A2"

    wb.save(path)
    return path
