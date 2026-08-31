"""Generate PRETAG AMIS seed data from the reviewed workbooks.
Outputs seed/zones.json, seed/districts.json, seed/district_aliases.json, seed/seed.sql
"""
import openpyxl, glob, os, json, unicodedata, re
from collections import OrderedDict

BASE = r"C:\Users\HP\PRETAG_R20_ANALYZER"
OUT  = os.path.join(BASE, "seed")
HEADERS = ["EMPLOYEE NO","NAME OF EMPLOYEE","MANAGEMENT UNIT","DISTRICT","REGION"]
os.makedirs(OUT, exist_ok=True)

def norm(s):
    if s is None: return ""
    return unicodedata.normalize("NFKC", str(s)).replace("\xa0"," ")
def clean(s):
    return re.sub(r"\s+"," ", norm(s)).strip()
def ncanon(s):
    return re.sub(r"[^a-z0-9]+"," ", norm(s).lower()).strip()
def code(s, n=3):
    letters = re.sub(r"[^A-Z]","", clean(s).upper())
    return (letters[:n] or "XXX")

ZONES = ["Adansi","Ahafo Ano","Amansie","Asante Akim","Asokore Mampong","Atwima Kwanwoma",
 "Atwima Nwabiagya","Bekwai","Bosome Freho","Bosomtwe","Ejisu","Ejura","Kumasi","Kwabre",
 "Mampong","Offinso","Sekyere Kumawu","Sekyere South & Central"]
FILEKEY = {z.upper().replace(" & "," AND "): z for z in ZONES}

# canonical district name cleanup: strip trailing " District"/" Municipal Assembly" noise words
def canon_district(sheetname):
    d = clean(sheetname).title()
    fix = {"Ejisu Municipal Assembly":"Ejisu Municipal",
           "Kwabre East Munnicipal":"Kwabre East Municipal",
           "Ejura-Sekyeredumase":"Ejura-Sekyeredumase Municipal",
           "Atwima Nwabiagya Municipal":"Atwima Nwabiagya Municipal"}
    if d in fix: return fix[d]
    # strip a trailing "District"/"Municipality" only if >=2 words remain
    stripped = re.sub(r"\s+(District|Municipality)$","", d)
    if len(stripped.split()) >= 2:
        d = stripped
    elif d.endswith(" District"):
        d = stripped  # e.g. "Akrofuom District" -> "Akrofuom"
    return d

districts = OrderedDict()      # canonical district -> zone
aliases   = {}                 # normalized_alias -> canonical district
alias_examples = {}            # canonical district -> set(raw alias strings)

for f in sorted(glob.glob(os.path.join(BASE,"R20 JULY","*.xlsx"))):
    key = clean(os.path.basename(f).split("-R20")[0]).upper()
    zone = FILEKEY[key]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows or [clean(c).upper() for c in rows[0][:5]] != HEADERS: continue
        sheet = ws.title
        cd = canon_district(sheet)
        # Regional Executive ruling: "Sekyere East" belongs to Ejisu Zone.
        if ncanon(cd).startswith("sekyere east"):
            cd, zone_use = "Sekyere East", "Ejisu"
        else:
            zone_use = zone
        districts.setdefault(cd, zone_use)
        for raw in [clean(sheet), cd] + [clean(r[3]) for r in rows[1:] if r and len(r)>3 and clean(r[3])]:
            aliases[ncanon(raw)] = cd
            alias_examples.setdefault(cd, set()).add(raw)
    wb.close()

# also fold in DISTRICT strings from both regional files
for rf in ["R20_Ashanti-6 July.xlsx","Ashanti-4 August.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(BASE,rf), read_only=True, data_only=True)
    ws = wb.active
    for r in ws.iter_rows(values_only=True):
        if not r or clean(r[0]).upper()=="EMPLOYEE NO": continue
        d = clean(r[3]) if len(r)>3 else ""
        if not d: continue
        c = ncanon(d)
        if c in aliases:
            alias_examples.setdefault(aliases[c], set()).add(d)
            continue
        # resolve by stripping noise words
        c2 = re.sub(r"\b(municipal|metropolitan|assembly|district|municipality)\b"," ", c).strip()
        c2 = re.sub(r"\s+"," ", c2)
        hit = None
        for k,cd in aliases.items():
            kk = re.sub(r"\b(municipal|metropolitan|assembly|district|municipality)\b"," ", k).strip()
            kk = re.sub(r"\s+"," ", kk)
            if kk and kk == c2: hit = cd; break
        if hit:
            aliases[c] = hit
            alias_examples.setdefault(hit, set()).add(d)
        else:
            print("!! UNRESOLVED regional district:", repr(d))
    wb.close()

# ---- build JSON ----
zones_json = [dict(id=i+1, zone_name=z, zone_code=code(z), region="Ashanti", is_active=True)
              for i,z in enumerate(ZONES)]
zid = {z["zone_name"]: z["id"] for z in zones_json}

used = set()
districts_json = []
for i,(d,z) in enumerate(sorted(districts.items())):
    base = re.sub(r"[^A-Z]","", clean(d).upper())[:4] or "DIST"
    cc, n = base, 1
    while cc in used:
        n += 1; cc = base[:3] + str(n)
    used.add(cc)
    districts_json.append(dict(id=i+1, district_name=d, district_code=cc, zone=z, zone_id=zid[z], is_active=True))
did = {d["district_name"]: d["id"] for d in districts_json}

aliases_json = []
seen = set()
for cd, ex in sorted(alias_examples.items()):
    for raw in sorted(ex):
        na = ncanon(raw)
        if na in seen: continue
        seen.add(na)
        aliases_json.append(dict(district=cd, district_id=did[cd], alias=raw, normalized_alias=na))

json.dump(zones_json, open(os.path.join(OUT,"zones.json"),"w"), indent=2)
json.dump(districts_json, open(os.path.join(OUT,"districts.json"),"w"), indent=2)
json.dump(aliases_json, open(os.path.join(OUT,"district_aliases.json"),"w"), indent=2, ensure_ascii=False)

# ---- SQL ----
def q(s): return "'" + str(s).replace("'","''") + "'"
lines = ["-- PRETAG AMIS seed data (generated from reviewed R20 workbooks)",
         "-- Sekyere East is assigned to Sekyere South & Central per Regional Executive ruling.",
         "", "insert into regions (region_name, region_code) values ('Ashanti','ASH');", ""]
lines.append("insert into zones (id, zone_name, zone_code, region_id, is_active) values")
lines.append(",\n".join(f"  ({z['id']}, {q(z['zone_name'])}, {q(z['zone_code'])}, 1, true)" for z in zones_json) + ";")
lines.append("")
lines.append("insert into districts (id, district_name, district_code, zone_id, is_active) values")
lines.append(",\n".join(f"  ({d['id']}, {q(d['district_name'])}, {q(d['district_code'])}, {d['zone_id']}, true)" for d in districts_json) + ";")
lines.append("")
lines.append("insert into district_aliases (district_id, alias, normalized_alias) values")
lines.append(",\n".join(f"  ({a['district_id']}, {q(a['alias'])}, {q(a['normalized_alias'])})" for a in aliases_json) + ";")
open(os.path.join(OUT,"seed.sql"),"w",encoding="utf-8").write("\n".join(lines)+"\n")

print(f"zones: {len(zones_json)}   districts: {len(districts_json)}   aliases: {len(aliases_json)}")
print("\nDistricts per zone:")
from collections import Counter
for z,c in Counter(d['zone'] for d in districts_json).most_common():
    print(f"  {z:26} {c}")
print("\nwrote:", ", ".join(sorted(os.listdir(OUT))))
